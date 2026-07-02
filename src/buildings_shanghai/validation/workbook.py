"""Build the annotation workbook (xlsx) for owner sign-off.

The layout mirrors the template the owner has already approved:

    编号 | 百度直达链接 (超链接) | 片区 | 抽样方式 | 建筑类型 (英文代码) |
    是否综合体 | 裙楼层数 | 总层数 | 层数把握 | 判断依据 | 街景拍摄年月 |
    备注 | lat_wgs84 | lon_wgs84

Blind-test rule: **no** column may contain a machine-derived height,
floor count, or archetype prediction. The annotator fills those in by
looking at the Baidu pin. See the module-level guard in
:func:`build_workbook`.

Dropdowns are wired via ``openpyxl``'s :class:`DataValidation` on named
lists. A second sheet ``类型代码对照`` gives the 14-way Chinese cheat
sheet.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


BUILDING_TYPE_CODES: list[tuple[str, str]] = [
    ("residential", "住宅"),
    ("office", "办公"),
    ("government_office", "政府办公"),
    ("hotel", "酒店"),
    ("shopping_mall", "商场/购物中心"),
    ("healthcare", "医院/医疗"),
    ("education", "学校/教育"),
    ("sports", "体育"),
    ("culture", "文化(博物馆/剧场/图书馆)"),
    ("transportation", "交通枢纽(车站/机场)"),
    ("exhibition", "会展"),
    ("mixed_use", "综合体(多功能)"),
    ("other_public", "其他公共"),
    ("unclear", "无法判断"),
]

MIXED_USE_OPTIONS: list[str] = ["是"]
CONFIDENCE_OPTIONS: list[str] = ["高", "中", "低"]
JUDGEMENT_OPTIONS: list[str] = ["招牌", "大门", "建筑外形", "地图标注", "卫星外形", "其他"]

_MAIN_HEADERS: list[str] = [
    "编号",
    "百度直达链接",
    "片区",
    "抽样方式",
    "建筑类型(英文代码)",
    "是否综合体",
    "裙楼层数",
    "总层数",
    "层数把握",
    "判断依据",
    "街景拍摄年月",
    "备注",
    "lat_wgs84(勿改)",
    "lon_wgs84(勿改)",
]

_BLIND_FORBIDDEN_COLS: frozenset[str] = frozenset(
    {"height", "height_m", "estimated_floor", "storeys_pred",
     "archetype", "archetype_pred", "predicted_type"}
)


class BlindTestViolationError(RuntimeError):
    """Raised when the input DataFrame carries a column the annotator
    must never see (per §5 blind-test rule)."""


def _check_blind_test(df: pd.DataFrame) -> None:
    leaked = _BLIND_FORBIDDEN_COLS & set(df.columns)
    if leaked:
        raise BlindTestViolationError(
            f"annotation workbook input carries forbidden columns: {sorted(leaked)}. "
            f"Blind-test rule §5: no height, no floor guess, no archetype prediction "
            f"may reach the annotator."
        )


def build_workbook(rows: pd.DataFrame, out_path: Path) -> Path:
    """Write ``annotation_workbook.xlsx`` at ``out_path``.

    ``rows`` must have exactly the columns:

        编号, 百度直达链接, 片区, 抽样方式, 备注, lat_wgs84, lon_wgs84

    The 8 annotator-filled columns are added blank; drop-down validation
    is attached to the four categorical ones plus 是否综合体.
    """
    _check_blind_test(rows)

    required = {"编号", "百度直达链接", "片区", "抽样方式", "备注", "lat_wgs84", "lon_wgs84"}
    missing = required - set(rows.columns)
    if missing:
        raise ValueError(f"workbook input missing columns: {sorted(missing)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "标注工作表"

    # Header row.
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="EEEEEE")
    for col_idx, h in enumerate(_MAIN_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    # Data rows.
    for i, row in rows.reset_index(drop=True).iterrows():
        r = i + 2  # 1-based, skipping header
        ws.cell(row=r, column=1, value=row["编号"])
        link_cell = ws.cell(row=r, column=2, value=row["百度直达链接"])
        link_cell.hyperlink = row["百度直达链接"]
        link_cell.font = Font(color="0563C1", underline="single")
        ws.cell(row=r, column=3, value=row["片区"])
        ws.cell(row=r, column=4, value=row["抽样方式"])
        # cols 5–11 left blank for the annotator
        ws.cell(row=r, column=12, value=row["备注"])
        ws.cell(row=r, column=13, value=float(row["lat_wgs84"]))
        ws.cell(row=r, column=14, value=float(row["lon_wgs84"]))

    # Column widths.
    widths = [8, 55, 10, 12, 20, 12, 10, 10, 10, 14, 14, 34, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _attach_dropdowns(ws, n_rows=len(rows))

    _add_lookup_sheet(wb)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _attach_dropdowns(ws, n_rows: int) -> None:
    last = n_rows + 1  # data rows: 2..last (inclusive)

    def _v(formula: str, prompt: str) -> DataValidation:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "请选择下拉框中列出的选项"
        dv.errorTitle = "无效输入"
        dv.prompt = prompt
        dv.showErrorMessage = True
        return dv

    # 建筑类型 — 14 items, inline list.
    type_codes = ",".join(code for code, _ in BUILDING_TYPE_CODES)
    dv_type = _v(f'"{type_codes}"', "选择 14 类之一;不确定时选 unclear")
    dv_type.error = (
        "请从下拉框选择,或若确需自定义值,先与 owner 讨论新增类别"
    )
    ws.add_data_validation(dv_type)
    dv_type.add(f"E2:E{last}")

    dv_mixed = _v(f'"{",".join(MIXED_USE_OPTIONS)}"', "只在建筑类型 = mixed_use 时才填 '是'")
    ws.add_data_validation(dv_mixed)
    dv_mixed.add(f"F2:F{last}")

    dv_conf = _v(f'"{",".join(CONFIDENCE_OPTIONS)}"', "对总层数的把握:高/中/低")
    ws.add_data_validation(dv_conf)
    dv_conf.add(f"I2:I{last}")

    # 判断依据 — allowUserInput=True so annotator can free-type "其他"变体.
    dv_judge = _v(
        f'"{",".join(JUDGEMENT_OPTIONS)}"',
        "选择主要判断线索;可从下拉选,也允许自填"
    )
    dv_judge.showErrorMessage = False  # allow free-typed values
    ws.add_data_validation(dv_judge)
    dv_judge.add(f"J2:J{last}")


def _add_lookup_sheet(wb: Workbook) -> None:
    """Add a "类型代码对照" sheet with the 14-way cheat sheet."""
    lookup = wb.create_sheet("类型代码对照")
    lookup.cell(row=1, column=1, value="英文代码").font = Font(bold=True)
    lookup.cell(row=1, column=2, value="中文含义").font = Font(bold=True)
    lookup.cell(row=1, column=3, value="备注").font = Font(bold=True)
    notes = {
        "residential": "含高层/多层住宅;公寓也归此",
        "office": "含金融/普通办公;不带政府背景",
        "government_office": "政府/机关办公",
        "hotel": "含酒店、宾馆、招待所",
        "shopping_mall": "综合市场也归此;不含单一便利店",
        "healthcare": "医院、卫生服务中心",
        "education": "含幼儿园—大学;宿舍归 residential",
        "sports": "体育场、健身房",
        "culture": "博物馆、剧场、图书馆",
        "transportation": "车站、机场航站楼",
        "exhibition": "会展中心",
        "mixed_use": "多功能综合体(需在'是否综合体'列填 是)",
        "other_public": "其他公共设施(如公厕、变电站)",
        "unclear": "任何一栏无把握时先勾此项",
    }
    for i, (code, zh) in enumerate(BUILDING_TYPE_CODES, start=2):
        lookup.cell(row=i, column=1, value=code)
        lookup.cell(row=i, column=2, value=zh)
        lookup.cell(row=i, column=3, value=notes.get(code, ""))
    lookup.column_dimensions["A"].width = 22
    lookup.column_dimensions["B"].width = 26
    lookup.column_dimensions["C"].width = 60
    lookup.freeze_panes = "A2"


__all__ = [
    "BUILDING_TYPE_CODES",
    "MIXED_USE_OPTIONS",
    "CONFIDENCE_OPTIONS",
    "JUDGEMENT_OPTIONS",
    "BlindTestViolationError",
    "build_workbook",
]
